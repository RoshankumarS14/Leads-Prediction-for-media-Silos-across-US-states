import os
import streamlit.components.v1 as components
import geopy.distance
import pandas as pd
import streamlit as st
from geopy.geocoders import Nominatim
import json
import plotly.express as px
import math
import numpy as np
from plot_utils import plot_gauge_Balance, plot_gauge_APScale
from plotly.subplots import make_subplots
from plotly.io import to_image
from openpyxl.drawing.image import Image as XLImage
from PIL import Image 
import io
from openpyxl import Workbook
from openpyxl import load_workbook
from map_utils import get_centre_zoom

st.set_page_config(
    page_title="Leads Prediction",
    page_icon="📈",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# URL of the map
map_url = "https://us-population-map.onrender.com/"

# Create a script that opens a new tab
st.markdown(f'<script>window.open("{map_url}", "_blank")</script>', unsafe_allow_html=True)

st.image("logo.png")

df_silo = pd.read_excel("Silo-Data-(12-04-23).xlsx")
role_adjuster = pd.read_excel("Silo-Data-Classifications.xlsx")
silos = df_silo["Silo"].unique()
ap_scale_silos = dict(df_silo[["Silo","AP-Scale"]].values)

if 'company_name' not in st.session_state:
    st.session_state.company_name = ''
if 'user_role' not in st.session_state:
    st.session_state.user_role = ''


col_name,col_name_input,col_role,col_role_input = st.columns([0.5,1.6,0.25,1.7]) 
col_name.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px;'>Company:</div>", unsafe_allow_html=True)
name_input_slot = col_name_input.empty()
st.session_state.company_name = name_input_slot.text_input('', '', key="Company_name")
col_role.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px;'>Role:</div>", unsafe_allow_html=True)
role_input_slot = col_role_input.empty()
st.session_state.user_role = role_input_slot.text_input('', '', key="role")
if len(st.session_state.user_role) > 18:
    st.warning(f"Input is too long! Please limit your input to 18 characters.")

cols_campaign = st.columns([0.7, 0.1, 0.9] * 3) 
campaigns=["Full:","Half:","Quarter:"]
campaigns_values = []
    
for j in range(3):

    col_name = cols_campaign[j*3]
    col_dollar_sign = cols_campaign[j*3 + 1]
    col_input = cols_campaign[j*3 + 2]
    
    # Display the option
    col_name.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px; margin-bottom:60px;'>{campaigns[j]}</div>", unsafe_allow_html=True)
    
    # Display the dollar sign
    col_dollar_sign.markdown(f"<div style='text-align: right; color: white; padding-top: 30px; font-size: 20px; margin-bottom:60px;'>$</div>", unsafe_allow_html=True)
    
    # Create the text input slot
    text_input_slot = col_input.empty()
    user_input = text_input_slot.text_input('', '', key=f'input_campaign{j}')
    campaigns_values.append(user_input)

role = st.selectbox("Select the job role:",role_adjuster["Role"],4)
# input_states = st.multiselect("Select the states:",states)
input_silos = st.multiselect("Select the Silos:",silos)
input_budget = []

st.text("Enter the budget for each of the media silos:")
# Calculate the number of rows
num_rows = len(input_silos) // 3
if len(input_silos) % 3 != 0:
    num_rows += 1

# For each row
for i in range(num_rows):
    # Create 9 columns: 3 for options, 3 for dollar signs, and 3 for inputs
    cols = st.columns([0.7, 0.1, 0.9] * 3)  # Repeat the pattern for 3 sets of columns
    
    # For each set of option, dollar sign, and input
    for j in range(3):
        # Calculate the index for the silo
        index = i * 3 + j
        
        # If the index is out of range, break the loop
        if index >= len(input_silos):
            break
        
        # Get the option
        option = input_silos[index]
        
        # Calculate the index for the columns
        col_option = cols[j*3]
        col_dollar = cols[j*3 + 1]
        col_input = cols[j*3 + 2]
        
        # Display the option
        col_option.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px;'>{option}</div>", unsafe_allow_html=True)
        
        # Display the dollar sign
        col_dollar.markdown(f"<div style='text-align: right; color: white; padding-top: 30px; font-size: 20px;'>$</div>", unsafe_allow_html=True)
        
        # Create the text input slot
        text_input_slot = col_input.empty()
        user_input = text_input_slot.text_input('', '', key=f'input_{index}')
        input_budget.append(user_input)


_RELEASE = False

if not _RELEASE:
    _component_func = components.declare_component(
        "my_component",
        url="https://us-population-map.onrender.com/",
    )
else:
    parent_dir = os.path.dirname(os.path.abspath(__file__))
    build_dir = os.path.join(parent_dir, "frontend/build")
    _component_func = components.declare_component("my_component", path=build_dir)

if "selected_states" not in st.session_state:
    st.session_state["selected_states"]=[]

if "rerun_flag" not in st.session_state:
    st.session_state["rerun_flag"]=False

def my_component(key=None):
    component_value = _component_func(key=key, default=0)
    return component_value

def get_state_name(lat, lon):
    geolocator = Nominatim(user_agent="my_geocoder")
    location = geolocator.reverse((lat, lon), exactly_one=True)
    address = location.raw['address']
    state = address.get('state', '')
    return state

def create_circle_feature(center, radius_miles, num_points=64, properties=None):
    """
    Create a GeoJSON feature representing a circle with a given center, radius, and number of points.
    :param center: Tuple of (longitude, latitude)
    :param radius: Radius in degrees (small values)
    :param num_points: Number of points to generate the circle
    :param properties: Dictionary of properties for the feature
    :return: GeoJSON feature dictionary
    """
    if properties is None:
        properties = {}

    # Earth's radius in miles
    earth_radius_miles = 3960.0

    # Calculate the radius in degrees latitude
    radius_lat = radius_miles / earth_radius_miles * (180 / math.pi)

    # Calculate the radius in degrees longitude, adjusting for the latitude
    radius_lng = radius_miles / (earth_radius_miles * math.cos(math.radians(center[1]))) * (180 / math.pi)

    # Apply the scale factor to the radius
    radius_lat *= 0.6
    radius_lng *= 0.6

    # Calculate the points of the circle
    circle_points = []
    for i in range(num_points):
        angle = math.radians(float(i) / num_points * 360)
        dx = radius_lng * math.cos(angle)
        dy = radius_lat * math.sin(angle)
        point = (center[0] + dx, center[1] + dy)
        circle_points.append(point)
    circle_points.append(circle_points[0])  # Ensure the polygon is closed

    # Create the GeoJSON feature
    feature = {
        "type": "Feature",
        "properties": properties,
        "geometry": {
            "type": "Polygon",
            "coordinates": [circle_points]
        }
    }
    return feature

# Define a function to calculate the distance between two points
def distance(point1, point2):
    return geopy.distance.distance(point1, point2).km

def refresh():
    st.experimental_rerun()

def calculate_rating(numbers):
    # Calculate the standard deviation
    std_dev = np.std(numbers)
    
    # Calculate the mean of the numbers
    mean = np.mean(numbers)
    
    # Normalize the standard deviation to a scale of 0 to 1
    normalized_std_dev = std_dev / mean if mean != 0 else 0
    
    # Calculate the rating
    rating = 100 * (1 - normalized_std_dev)
    
    # Make sure the rating is within the scale of 1 to 100
    rating = max(min(rating, 100), 1)
    
    return rating

# URL of the map
map_url = "https://us-population-map.onrender.com/"
 
if not _RELEASE:

    col_head,col_but = st.columns([2,1])
    with col_head:
        st.subheader("Population Map - US")
    with col_but:
        # Create a button that opens the link in a new tab
        st.markdown(f'<a href="{map_url}" target="_blank"><input type="button" value="Open Map" style="color: white; background-color: #FF4B4B; border: none; border-radius: 5px; padding: 10px 20px;"></a>', unsafe_allow_html=True)
   
    clicked_coords = my_component()
 
    # st.markdown(clicked_coords)
    if not (type(clicked_coords)==list):
        # st.write(clicked_coords.get("lat"), clicked_coords.get("lng"))
        try:
            selected_state = get_state_name(clicked_coords.get("lat"), clicked_coords.get("lng")) 
            if selected_state not in st.session_state["selected_states"]:
                st.session_state["selected_states"].append(selected_state)
        except:
            pass    

# Read the csv file into a pandas dataframe
data = pd.read_csv("US_Population.csv")

# US state names and abbreviations
us_state_abbreviations = {
    'Alabama': 'AL','Alaska': 'AK','Arizona': 'AZ','Arkansas': 'AR','California': 'CA','Colorado': 'CO','Connecticut': 'CT','Delaware': 'DE',
    'Florida': 'FL','Georgia': 'GA','Hawaii': 'HI','Idaho': 'ID','Illinois': 'IL','Indiana': 'IN','Iowa': 'IA','Kansas': 'KS','Kentucky': 'KY',
    'Louisiana': 'LA','Maine': 'ME','Maryland': 'MD','Massachusetts': 'MA','Michigan': 'MI','Minnesota': 'MN','Mississippi': 'MS','Missouri': 'MO',
    'Montana': 'MT','Nebraska': 'NE','Nevada': 'NV','New Hampshire': 'NH','New Jersey': 'NJ','New Mexico': 'NM','New York': 'NY',
    'North Carolina': 'NC','North Dakota': 'ND','Ohio': 'OH','Oklahoma': 'OK','Oregon': 'OR','Pennsylvania': 'PA','Rhode Island': 'RI',
    'South Carolina': 'SC','South Dakota': 'SD','Tennessee': 'TN','Texas': 'TX','Utah': 'UT','Vermont': 'VT','Virginia': 'VA',
    'Washington': 'WA','West Virginia': 'WV','Wisconsin': 'WI','Wyoming': 'WY'
}

# Load GeoJSON file
with open('gz_2010_us_040_00_500k.json') as response:
    states_json = json.load(response)

if "states" not in st.session_state:
    st.session_state["states"]=[]

if "states_json" not in st.session_state:
    st.session_state["states_json"]=states_json

st.session_state["states"] = st.multiselect("Selected States:",st.session_state["selected_states"],st.session_state["selected_states"])

st.session_state["rerun_flag"]=False

calculate = st.button("Predict!",use_container_width=True)

def set_ad_campaign():
    st.session_state.ad_campaign = True

def set_predict_leads():
    st.session_state.predict_leads = True

if 'ad_campaign' not in st.session_state:
    st.session_state.ad_campaign = False

if 'predict_leads' not in st.session_state:
    st.session_state.predict_leads = False

if "state_wise_pop" not in st.session_state:
    st.session_state["state_wise_pop"] = pd.DataFrame()

if "state_df" not in st.session_state:
    st.session_state["state_df"] = pd.DataFrame()

if calculate:

    dfs = []
    circle_count=0
    circles = []
    if type(clicked_coords)==list:
        for coords in clicked_coords:
            center = coords.get("center")
            radius = coords.get("radius")/1000

            # Filter the dataframe to keep only the cities that are within the radius
            df_filtered = data[data.apply(lambda row: distance((center["lat"],center["lng"]), (row["lat"], row["lon"])) <= radius, axis=1)]
            dfs.append(df_filtered)

            circle_count+=1
            center = (center["lng"],center["lat"])
            print(center,radius)
       
            # Define the properties for the new circular state
            new_state_properties = {
                'NAME': 'CircularState'+str(circle_count),  # The name of the new state
            }

            # Create the circular state feature with the updated properties
            new_state_feature = create_circle_feature(
                center=center,
                radius_miles=radius,
                num_points=200,  # More points for a smoother circle
                properties=new_state_properties
            )

            # Add the new state to the existing GeoJSON data
            st.session_state["states_json"]['features'].append(new_state_feature)
            circles.append("CircularState"+str(circle_count))
            

    selected_states = [us_state_abbreviations[state] for state in st.session_state["states"]]
    if len(selected_states)>=1:
        dfs.append(data[data["state"].isin(selected_states)])
    
    
    df = pd.concat(dfs,ignore_index=True)
    df.drop_duplicates(inplace=True)
    # Calculate the total population of the filtered cities
    total_population = df["population"].sum()
    
    # Display the result
    # st.write(f"The total population of the cities within {radius} km of {center} is {total_population}.")
    st.write(f"The total population of the cities within the circle is {total_population}.")

    state_wise_pop = pd.pivot_table(data=df,index="state",values="population",aggfunc=sum)
    state_wise_pop["percentage"] = state_wise_pop["population"]/total_population*100
    state_wise_pop["percentage"] = state_wise_pop["percentage"].apply(lambda a : round(a,2))
    state_wise_pop.reset_index(inplace=True)
    state_wise_pop.columns = ["State","Population","Percentage Population"] 

    st.session_state["state_wise_pop"] = state_wise_pop
    df_html = state_wise_pop.to_html(classes='table table-striped',index=False)
    df_html = df_html.replace('<table ','<table style="text-align:right; margin-bottom:40px; margin-top:50px; width:95%;" ')

    st.session_state["states"].extend(circles)
    
    # Create a DataFrame with the states to highlight
    state_df = pd.DataFrame({
        'state': st.session_state["states"],  # replace with your states
        'highlight': [1 for i in st.session_state["states"]]  # 1 means highlight
    })
    st.session_state["state_df"] = state_df

    center_lat,center_lon,zoom_level = get_centre_zoom(st.session_state["states_json"],state_df["state"])
    
    # Create a choropleth map
    fig = px.choropleth_mapbox(state_df, geojson=st.session_state["states_json"], locations='state', color='highlight',
                        color_continuous_scale=["rgba(0,0,255,0.1)", "rgba(0,0,255,0.5)"], range_color=(0, 1),
                        labels={'highlight':'highlight'},
                        featureidkey="properties.NAME")
    
    # Update the layout
    fig.update_geos(showcountries=False, showcoastlines=True, showland=True, fitbounds="locations")
    fig.update_layout(mapbox_style="mapbox://styles/mapbox/streets-v11", 
                    mapbox_zoom=zoom_level, 
                    mapbox_accesstoken ="pk.eyJ1IjoidGpkMjAyNCIsImEiOiJjbHIyNmU2Z2oweTRmMnFuMWN1dmN1N3V4In0.9dU2pjRURk4qs31aBAV4lg",
                    mapbox_center = {"lat": center_lat, "lon": center_lon},
                    coloraxis=dict(showscale=False))
    # fig.update_layout(height=300, margin={"r":0,"t":0,"l":0,"b":0})

    c1,c2 = st.columns([1.3,2])
    
    with c1:
        st.markdown(df_html, unsafe_allow_html=True)

    with c2:
        # Display the map in Streamlit
        st.plotly_chart(fig,use_container_width=True)

#     st.button("Predict", on_click=set_predict_leads, use_container_width=True)

# if st.session_state.predict_leads:
    leads = []
    input_budget = [float(i) for i in input_budget]
    state_wise_pop = st.session_state["state_wise_pop"]
    for silo,budget in zip(input_silos,input_budget):
        lead = 0
        for state,percent in zip(state_wise_pop["State"],state_wise_pop["Percentage Population"]):
            state_budget = percent*budget/100
            if state in df_silo["ST"]:
                lead += state_budget/df_silo[(df_silo["Silo"]==silo) & (df_silo["ST"]==state)]["CPL"].values[0]
            else:
                lead += state_budget/df_silo[(df_silo["Silo"]==silo) & (df_silo["ST"]=="US")]["CPL"].values[0]
        leads.append(lead)
    AP_scales = [ap_scale_silos[silo]*lead for silo,lead in zip(input_silos,leads)]
    average_AP_scales = round(sum(AP_scales)/sum(leads),1)
    leads = [round(i,1) for i in leads]

    result = pd.DataFrame({"Silo":input_silos,"Budget":input_budget,"Leads":leads})
    result["Budget"] = result["Budget"].apply(lambda a: "$ " + '{:.2f}'.format(a))
    result.index = np.arange(1,len(result)+1)
    # st.dataframe(result)

    # Convert the DataFrame to HTML and align all columns to the right
    df_html = result.to_html(classes='table table-striped')
    df_html = df_html.replace('<table ','<table style="text-align:right; margin-bottom:40px; margin-top:50px; width:95%;" ')

    col_df,col_gauge = st.columns([1.5,1])

    with col_df:
        # Display the DataFrame
        st.markdown(df_html, unsafe_allow_html=True)
        # st.write("AP Scale: "+str(average_AP_scales))
        # st.write("Total Budget: $"+str(sum(input_budget)))
        # st.write("Target Leads: "+str(round(sum(leads))*2))
        # st.write("Min Leads: "+str(round(sum(leads))))
    
        adjuster = role_adjuster[role_adjuster["Role"]==role]["Adjuster"].values[0]
        # Create formatted strings

        col_result1,col_result2 = st.columns([1,1])

        with col_result1:
            st.write(f"""
            <pre>
            <table style="border: none;">
            <tr><td style="text-align: left; border: none;">AP Scale:</td><td style="text-align: right; border: none;">{int(average_AP_scales*10)}</td></tr>
            <tr><td style="text-align: left; border: none;">Balance:</td><td style="text-align: right; border: none;">{int(calculate_rating(np.log([10000 if i>10000 else i for i in input_budget if i>50])))}</td></tr>
            <tr><td style="text-align: left; border: none;">Overhead Funds:</td><td style="text-align: right; border: none;">{round((1-(sum(input_budget)/float(campaigns_values[0])))*100)}</td></tr>
            </table>
            </pre>
            """, unsafe_allow_html=True)

        with col_result2:
            st.write(f"""
            <pre>
            <table style="border: none;">
            
            <tr><td style="text-align: left; border: none;">Target Leads:</td><td style="text-align: right; border: none;">{round(sum(leads)*adjuster)*2}</td></tr>
            <tr><td style="text-align: left; border: none;">Min Leads:</td><td style="text-align: right; border: none;">{round(sum(leads)*adjuster)}</td></tr>
            <tr><td style="text-align: left; border: none;">Min Budget:</td><td style="text-align: right; border: none;">${str(int(sum(input_budget)))}</td></tr>
            </table>
            </table>
            </pre>
            """, unsafe_allow_html=True)

    with col_gauge:
        trace1 = plot_gauge_APScale(int(average_AP_scales*10))
        trace2 = plot_gauge_Balance(int(calculate_rating(np.log([10000 if i>10000 else i for i in input_budget if i>50]))))
        # Create a subplot and add the gauges to it
        fig = make_subplots(rows=2, cols=1, specs=[[{'type': 'indicator'}],[{'type': 'indicator'}]],vertical_spacing=0.2)
        fig.append_trace(trace1, row=1, col=1)
        fig.append_trace(trace2, row=2, col=1)
        
        # Display the chart in Streamlit
        st.plotly_chart(fig, use_container_width=True)

    trace3 = plot_gauge_APScale(int(average_AP_scales*10),title="")
    fig2 = make_subplots(rows=1, cols=1, specs=[[{'type': 'indicator'}]])
    fig2.append_trace(trace3, row=1, col=1)

    fig2_bytes = to_image(fig2, format="png")
    img = Image.open(io.BytesIO(fig2_bytes))
    # Resize the image
    width, height = img.size
    new_width = 360
    new_height = 155
    img = img.resize((new_width, new_height))
    # Save the resized image to a BytesIO object
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr = img_byte_arr.getvalue()
    
    # Load the workbook
    wb = load_workbook('New-Template.xlsx')
    # Select the sheet
    sheet = wb['juliabid'] 
    
    # Create an Image object from BytesIO object
    img = XLImage(io.BytesIO(img_byte_arr))
    # Add the image to the sheet
    sheet.add_image(img, 'B36')

    state_df = st.session_state["state_df"]
    fig3 = px.choropleth_mapbox(state_df, geojson=st.session_state["states_json"], locations='state', color='highlight',
                        color_continuous_scale=["rgba(0,0,255,0.1)", "rgba(0,0,255,0.5)"], range_color=(0, 1),
                        labels={'highlight':'highlight'},
                        featureidkey="properties.NAME")
    
    # Update the layout
    fig3.update_geos(showcountries=False, showcoastlines=True, showland=True, fitbounds="locations")
    fig3.update_layout(mapbox_style="mapbox://styles/mapbox/streets-v11", 
                    mapbox_zoom=zoom_level, 
                    mapbox_accesstoken ="pk.eyJ1IjoidGpkMjAyNCIsImEiOiJjbHIyNmU2Z2oweTRmMnFuMWN1dmN1N3V4In0.9dU2pjRURk4qs31aBAV4lg",
                    mapbox_center = {"lat": center_lat, "lon": center_lon},
                    coloraxis=dict(showscale=False))
    fig3.update_layout(margin={"r":0,"t":0,"l":0,"b":0})
    
    fig3_bytes = to_image(fig3, format="png")
    map_img = Image.open(io.BytesIO(fig3_bytes))
    # Resize the image
    width, height = map_img.size
    new_width = 360
    new_height = 310
    map_img = map_img.resize((new_width, new_height))

    # Save the resized image to a BytesIO object
    map_img_byte_arr = io.BytesIO()
    map_img.save(map_img_byte_arr, format='PNG')
    map_img_byte_arr = map_img_byte_arr.getvalue()

    # Create an Image object from BytesIO object
    map_img = XLImage(io.BytesIO(map_img_byte_arr))
    # Add the image to the sheet
    sheet.add_image(map_img, 'B20')

    # Write DataFrame to Excel from cell X11 for the first column
    for i, value in enumerate(result.iloc[:, 0]):
        sheet.cell(row=i+11, column=24, value=value)
    
    # Write DataFrame to Excel from cell Y11 for the second column
    for i, value in enumerate(input_budget):
        sheet.cell(row=i+11, column=25, value=value)
    
    # Write DataFrame to Excel from cell AA11 for the third column
    for i, value in enumerate(leads):
        sheet.cell(row=i+11, column=27, value=value)
        
    campaigns_values = [0 if i=="" else float(i) for i in campaigns_values] 
    # Write DataFrame to Excel from cell Y6 for the campaigns_values
    for i, value in enumerate(campaigns_values):
        sheet.cell(row=i+6, column=25, value=value if value != "" else 0.0)

    sheet.cell(row=5,column=25,value=st.session_state.company_name)
    sheet.cell(row=7,column=27,value=st.session_state.user_role)
    
    # Save the workbook to a BytesIO object
    excel_byte_arr = io.BytesIO()
    wb.save(excel_byte_arr)
        
    file_name = "TJD-" + st.session_state.company_name + ".xlsx"
    with open("New-Template.xlsx", "rb") as file:
         file_bytes = file.read()
    if st.download_button(
        label="Create Campaign!",
        data=excel_byte_arr.getvalue(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        ):
            st.session_state["rerun_flag"]=True

if st.session_state["rerun_flag"]:
    st.experimental_rerun()




    
